from flask import Flask, request, render_template
import cv2
import os

import openpyxl

app = Flask(__name__)
desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop\\')
app.config['UPLOAD_FOLDER'] = desktop


@app.route('/')
def default():
    return render_template("index.html")

@app.route('/sales.html')
def sales():
    return render_template("sales.html")

@app.route('/predict')
def predict():
    if not (os.path.exists(f'{desktop}/data.csv')):
        return 'Data file does not exist. Please enter data first'

@app.route('/form1data', methods =['GET','POST'])
def form1():
    date = request.form['date']
    itemid = request.form['itemid']
    quantiy = request.form['quantity']

    if not (os.path.exists(f'{desktop}/data.csv')):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet.cell(row=1, column=1, value="id")
        sheet.cell(row=1, column=2, value="date")
        sheet.cell(row=1, column=3, value="store")
        sheet.cell(row=1, column=4, value="item")
        sheet.cell(row=1, column=5, value="sales")
        
        sheet.cell(row=2, column=1, value=0)
        sheet.cell(row=2, column=2, value=date)
        sheet.cell(row=2, column=3, value=1)
        sheet.cell(row=2, column=4, value=itemid)
        sheet.cell(row=2, column=5, value=quantiy)
        
        


    #return  str(type(date))+ '  ' + str(type(itemid)) + '  ' + str(type(quantiy))
    

if __name__ == "__main__":
    app.run(debug=True)