import numpy as np
from datetime import date, timedelta
import openpyxl
year = 3
current_date = np.datetime64(date.today())
end_date = current_date + np.timedelta64(365*year, 'D')
dates = np.arange(current_date, end_date, np.timedelta64(1, 'D'))
formatted_dates = np.array([np.datetime_as_string(date, unit='D') for date in dates])

formatted_dates = np.tile(formatted_dates, 50)
store = np.ones(365*50*year)
item = np.repeat(np.arange(1, 51), 365*year)
print(item)
print(len(formatted_dates), len(store), len(item))

workbook = openpyxl.Workbook()
sheet = workbook.active

sheet.cell(row=1, column=1, value="Date")
sheet.cell(row=1, column=2, value="Store")
sheet.cell(row=1, column=3, value="Item")

# Write data to the columns
for i in range(len(formatted_dates)):
    sheet.cell(row=i+2, column=1, value=formatted_dates[i])
    sheet.cell(row=i+2, column=2, value=store[i])
    sheet.cell(row=i+2, column=3, value=item[i])

# Save the workbook to a file
workbook.save('train_3yr.csv')