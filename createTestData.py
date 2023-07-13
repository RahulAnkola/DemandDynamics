import numpy as np
from datetime import datetime, timedelta
import openpyxl

start_date = np.datetime64('2026-07-09')
end_date = start_date + np.timedelta64(90, 'D')
dates = np.arange(start_date, end_date, np.timedelta64(1, 'D'))
formatted_dates = np.datetime_as_string(np.repeat(dates, 50), unit='D')

id = np.arange(4500)
store = np.ones(4500)
item = np.repeat(np.arange(1, 51), 90)

workbook = openpyxl.Workbook()
sheet = workbook.active

sheet.cell(row=1, column=1, value="id")
sheet.cell(row=1, column=2, value="date")
sheet.cell(row=1, column=3, value="store")
sheet.cell(row=1, column=4, value="item")

# Write data to the columns
for i in range(len(formatted_dates)):
    sheet.cell(row=i+2, column=1, value=id[i])
    sheet.cell(row=i+2, column=2, value=formatted_dates[i])
    sheet.cell(row=i+2, column=3, value=store[i])
    sheet.cell(row=i+2, column=4, value=item[i])

# Save the workbook to a file
workbook.save('test_3mths.csv')
